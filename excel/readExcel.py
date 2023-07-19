import json

from openpyxl import load_workbook


class Excel:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.sheet_name = None

    def read_case_excel(self, sheet_name):
        '''
        读取测试用例文件
        :param sheet_name:
        :return: 测试用例列表，每个元素为一个字典，字典的键为表头，值为单元格的值
        '''
        self.sheet_name = sheet_name
        try:
            # 加载现有的工作簿
            excel_obj = load_workbook(self.excel_path)
            # 获取指定名称的工作表
            ws = excel_obj[sheet_name]
        except Exception as e:
            print(f'读取测试用例文件失败：{e}')
            return
        # 读取表头作为字典的键
        headers = [cell.value for cell in ws[1]]
        # 读取数据
        data = []
        # ws.iter_rows(min_row=2) 方法用于遍历工作表中的每一行。其中，min_row=2 参数表示从第二行开始遍历。
        # values_only：指定是否只返回单元格中的值。如果设置为 True，则该方法将只返回单元格中的值；否则，它将返回单元格对象。
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for header, value in zip(headers, row):
                record[header] = value
            data.append(record)
        return data

    def read_conf_excel(self, sheet_name='config'):
        '''
        读取配置文件
        :param sheet_name: 配置文件工作表名称
        :return: 配置数据字典
        '''
        self.sheet_name = sheet_name
        try:
            # 加载现有的工作簿
            excel_obj = load_workbook(self.excel_path)
            # 获取指定名称的工作表
            ws = excel_obj[sheet_name]
        except Exception as e:
            print(f'读取配置文件失败：{e}')
            print(e)
            return
        # 存放配置数据的字典
        data = {}
        for row in ws.iter_rows(max_col=2, values_only=True):
            # any 是一个内置函数，用于检查可迭代对象中是否存在至少一个元素为真。
            if not any(row):
                continue
            key, value = row
            if key == 'appiumcaps':
                value = json.loads(value)
            data[key] = value
        return data

def get_all_test_sheet_name(excel_path):
    '''
    获取所有测试用例工作表名称
    :return: 返回所有以test开头或者以test结尾的工作表名称
    '''
    excel_obj = load_workbook(excel_path)
    test_sheet_names = [i for i in excel_obj.sheetnames if i.startswith("test") or i.endswith("test")]
    return test_sheet_names


if __name__ == '__main__':
    path = './data/data.xlsx'
    excel = Excel(path)
    print(get_all_test_sheet_name(path))
    print(excel.read_case_excel('test_dsw_app_sc_037'))
    print(excel.read_conf_excel())
    print(excel.read_conf_excel().get('appiumcaps').get('appPackage'))
